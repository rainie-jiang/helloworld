# print("rainie jiang")
# print("*" * 10 )
# name = "John Smith"
# age = 20
# new_patient = True
# # print(name, age, new_patient)
# input("what is your name")
# name = input("what is your name? ")
# # print("hi " + name)
# # name=input("what is your name? ")
# # favorite_color=input("what is your favorite color ")
# # print(name + " like "+ favorite_color)
# weight=input("what is your weight in pounds? ")
# # print(int(weight)*0.45)
# print("rainie jiang".upper())
# buyer_has_good_credit=False
# price_of_a_house=1000000
#
# if buyer_has_good_credit:
#     down_payment=price_of_a_house*0.1
#     print(down_payment)
# else:
#     down_payment=price_of_a_house*0.2
#     print(down_payment)
# print(f"Down payment: {down_payment}")
# name=input("what is your name?")
# name_length=len(name)
# if name_length<3:
#     print("name must be at least 3 characters")
# elif name_length>50:
#     print("name can be a maximum of 50 characters")
# else:
# # LOOP Syntax练习
# command=""
# started= False
# while True:
#     command=input("> ").lower()
#     if command=="help":
#         print('''start-to start the car
#         stop-to stop the car
#         quit-to exit
#         ''')
#     elif command=="start":
#         if started==False:
#             print("car started...Ready to go!")
#             started=True
#         else:
#             print("car already started")
#     elif command=="stop":
#         if started==False:
#             print("car already stopped")
#         else:
#             started = False
#             print("car stopped")
#     elif command == "quit":
#         break
#     else:
#         print("i don't understand that...")
# #
# total=0
# for item in [10,20,30]:
#     print(item)
# #     total=total+item
# # print(total)
# list=[5,2,5,2,2]
# for item in list:
# #     print("*"*item)
# output=""
# for item in range(5):
# #     output=output+"X"
# # print(output)
# list=[3,5,9,12]
# max=0
# for item in list:
#     if max<item:
#         max=item
# print(max)
# matrix= [
#     [1,2,3],
#     [4,5,6],
#     [7,8,9]
# ]
# for row in matrix:
#     for item in row:
#         print(item)
# LIst
# number=[5,2,1,7,4]
# number.append(20)
# print(number)

# numbers=[5,2,1,7,4]
# numbers.insert(0,10)
# print(numbers)

# numbers=[5,2,1,7,4]
# numbers.remove(2)
# print(numbers)

# numbers=[5,2,1,7,4]
# numbers.clear()
# print(numbers)

# numbers=[5,2,1,7,4]
# numbers.pop()
# print(numbers)

# numbers=[5,2,1,7,4]
# print(numbers.index(5))

# numbers=[5,2,1,7,4]
# print(50 in numbers)

# numbers=[5,2,1,5,7,4]
# print(numbers.count(5))

# numbers=[5,2,1,5,7,4]
# print(numbers.sort())
# print(numbers)

# numbers=[5,2,1,5,7,4]
# numbers.reverse()
# print(numbers)

# numbers=[5,2,1,5,7,4]
# numbers2=numbers.copy()
# numbers.append(10)
# print(numbers2)

# numbers=[2,2,4,6,3,4,6,1]
# uniques=[]
# for number in numbers:
#     if number not in uniques:
#         uniques.append(number)
# print(uniques)

# tuple: (1,2,3)
# list: [1,2,3]
#
# Unpacking
# coordinates=(1,2,3)
# x = coordinates[0]
# y = coordinates[1]
# z = coordinates[2]
# x,y,z = coordinates
# print(x)
# print(y)
# print(z)

# Dictionaries
# customer= {
#     "name": "John Smith",
#     "age": 30,
#     "is_verified": True
# }
# print(customer["name"])
# print(customer.get("name"))
# print(customer.get("birthdate","Jan 1 1980"))
# customer["name"]="Jack Smith"
# print(customer["name"])
# customer["birthdate"]="Jan 1 1980"
# print(customer["birthdate"])

# phones = {
#     "1": "one",
#     "2": "two",
#     "3": "three"
# }
# new_phone=input("what is your phone number?" )
# for x in new_phone:
#     print(x)
#     print(phones.get(x))
# !!!!!!!!!
# phone = input("phone: ")
# digits_mapping = {
#     "1": "one",
#     "2": "two",
#     "3": "three",
#     "4": "four"
# }
# output=""
# for ch in phone:
#     output += digits_mapping.get(ch,"!") + " "
# print(output)

# Emoji Converter!!!!!!
# message = input(">")
# words = message.split(" ")
# emojis = {
#     ":)": "🙂",
#     ":(": "😔"
# }
# output=""
# for word in words:
#     output += emojis.get(word, word) + " "
# print(output)

# def greet_user():
#     print("Hi there!")
#     print("welcome aboard")
#
#
# print("start")
# greet_user()
# print("finish")

# Parameters
# def greet_user(name):
#     print(f"Hi {name}!")
#     print("welcome aboard")
#
#
# print("start")
# greet_user("John")
# print("finish")

# def greet_user(name):
#     print(f"Hi {name}!")
#     print("welcome aboard")
#
#
# print("start")
# greet_user("John")
# greet_user("Mary")
# print("finish")

# def greet_user(first_name, last_name):
#     print(f"Hi {first_name} {last_name}!")
#     print("welcome aboard")
#
#
# print("start")
# greet_user("John","Smith")
# print("finish")

# KEYWORD ARGUMENTS实参
# Return Statement

# def square(number):
#     return number * number
#
#
# result=square(3)
# print(result)
# print(square(3))

# y=f(x)


# def add(a, b):
#     # print(a + b)
#     return a + b
# output = 0
# output = add(1, 2)
# print('output', output)

# def multiply(a,b):
#     return a*b
#
#
# print('output', multiply(2,3))

# creating a reusable function
#
# def emoji_converter(message):
#     words = message.split(" ")
#     emojis = {
#         ":)": "🙂",
#         ":(": "😔"
#     }
#     output = ""
#     for word in words:
#         output += emojis.get(word, word) + " "
#     return output
#
#
# message = input(">")
# result= emoji_converter(message)
# print(result)

# Exception(排除错误）
# try:
#     age = int(input("age: "))
#     print(age)
# except ValueError:
#     print("Invalid value")
#
# try:
#     age = int(input("age: "))
#     income = 2000
#     risk = income / age
#     print(age)
# except ZeroDivisionError:
#     print("Age cannot be 0.")
# except ValueError:
#     print("Invalid value")

# COMMENT
# CLASS面向对象：例如numbers.string.booleans. lists. dictionaries
# class  Point:
#     def move(self):
#         print("move")
#
#     def draw(self):
#         print("draw")
#
#
# point1= Point()
# point1.x= 10
# point1.y= 20
# print(point1.x)
# point1.draw()
# Point.draw(point1)

# Constructors构造器
# （)的是函数


# class Point:
#     def __init__(self, x, y):
#         self.x = x
#         self.y = y
#
#     def move(self):
#         print("move")
#
#     def draw(self):
#         print("draw")
#
#
# point = Point(10, 20)
# point.x= 11
# print(point.x)

# class Person:
#     def __init__(self, name):
#         self.name=name
#
#     def talk(self):
#         print("talk")
#
#
# person=Person("Jack")
# print(person.name)
# person.talk()
# print(person.talk)
# 函数可以直接run ,例如print

# class Person:
#     def __init__(self, name):
#         self.name=name
#
#     def talk(self):
#         print(f"Hi, I am {self.name}")
#
#
# person=Person("Jack")
# print(person.name)
# person.talk()
# person2=Person("Ryan")
# person2.talk()

# Inheritance继承
# class Mammal:
#     def walk(self):
#         print("walk")
#
#
# class Dog(Mammal):
#     def bark(self):
#         print("bark")
#
#
# class Cat(Mammal):
#     pass
#
#
# dog1=Dog()
# dog1.name="Ryan"
# print(dog1.name)
# dog1.walk()
# dog1.bark()

# Modules模块-import(导入不同模块文件夹的命令语句）
# def lbs_to_kg(weight):
#     return weight * 0.45
#
#
# def kg_to_lbs(weight):
#     return weight / 0.45

# numbers = [10,3,6,2]
# max = numbers[0]
# for number in numbers:
#     if number > max:
#         max = number
# print(max)

# import utils
# # from utils import find_max
# list=[3,6,9,12]
# # print(utils.find_max(list))
# print(max(list))

# packages 包裹
# generating random values
# import random

# for i in range(3):
#     print(random.randint(10,20))
#
# members = ["john", "mary", "bob", "mosh"]
# leader= random.choice(members)
# print(leader)

# import dice
# print(dice.roll())
# import random
#
#
# class Dice:
#     def roll(self):
#         first = random.randint(1,6)
#         second = random.randint(1,6)
#         return first, second
# dice = Dice()
# print(dice.roll())

# files and directories
# from pathlib import Path

#class的首字母要upperclass, function函数不需要且加（）

# path= Path("ecommerce")
# print(path.exists())
# path= Path("emails")
# # print(path.mkdir())
# print(path.rmdir())

# from pathlib import Path
# path= Path("")
# # print(path.glob("*.py"))
# for file in path.glob("*.py"):
#     print(file)
#
# from pathlib import Path
# path= Path("")
# for file in path.glob("*"):
#     print(file)

#Pypi and pip模块
# dictionary:字典
# directory: 文件夹
# modules模块》-package包
# coordinate:坐标

# project1: Excel Spreadsheets (3:56)
# import openpyxl as xl
# wb=xl.load_workbook("transactions.xlsx")
# sheet= wb["Sheet1"]
# cell=sheet["a1"]
# cell=sheet.cell(1,1)
# print(sheet.max_row)

# import openpyxl as xl
# wb=xl.load_workbook("transactions.xlsx")
# sheet= wb["Sheet1"]
# cell=sheet["a1"]
# cell=sheet.cell(1,1)
#
# for row in range(1, sheet.max_row + 1):
#     print(row)

# 给excel增加一列带有90%折扣价格的列
# import openpyxl as xl
# wb=xl.load_workbook("transactions.xlsx")
# sheet= wb["Sheet1"]
# cell=sheet["a1"]
# cell=sheet.cell(1,1)
#
# for row in range(2, sheet.max_row + 1):
#     cell=sheet.cell(row,3)
#     print(cell.value)
#     corrected_price = cell.value*0.9
#     corrected_price_cell = sheet.cell(row,4)
#     corrected_price_cell.value=corrected_price
#
# wb.save("transactions2.xlsx")

# 给excel的value值增加chart
# import openpyxl as xl
# from openpyxl.chart import BarChart, Reference
#
# wb=xl.load_workbook("transactions.xlsx")
# sheet= wb["Sheet1"]
# cell=sheet["a1"]
# cell=sheet.cell(1,1)
#
# for row in range(2, sheet.max_row + 1):
#     cell=sheet.cell(row,3)
#     print(cell.value)
#     corrected_price = cell.value*0.9
#     corrected_price_cell = sheet.cell(row,4)
#     corrected_price_cell.value=corrected_price
#
# values = Reference(sheet,
#           min_row=2,
#           max_row=sheet.max_row,
#           min_col=4,
#           max_col=4)
#
# chart = BarChart()
# chart.add_data(values)
# sheet.add_chart(chart, "e2")
#
# wb.save("transactions2.xlsx")

# 让上面的syntax变得更加organized和def函数化
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):

    wb=xl.load_workbook(filename)
    sheet= wb["Sheet1"]


    for row in range(2, sheet.max_row + 1):
        cell=sheet.cell(row,3)
        print(cell.value)
        corrected_price = cell.value*0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value=corrected_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")

    wb.save(filename)